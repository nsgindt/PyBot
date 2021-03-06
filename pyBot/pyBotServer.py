import time, datetime, subprocess, sys, threading, json, os.path
#import logging
import pyautogui as botgui
import helper_library as helper
import math
from openpyxl import load_workbook
from flask import Flask, render_template, request, session
from flask_socketio import SocketIO, emit, disconnect


#logging.basicConfig(level=logging.DEBUG, format='[%(levelname)s] (%(threadName)-10s) %(message)s',)
                    	
# async setting 
async_mode = 'eventlet'

#setting config file path
my_path = os.path.dirname(__file__)
config_path = os.path.join(my_path, 'config/bot_config.json')

app = Flask(__name__)
socketio = SocketIO(app, async_mode=async_mode)
thread_stop_event = threading.Event()


@socketio.on('connect')
def __on_connect():
	print('New connection is established')

#page routing
@app.route("/")
def main():
	with open(config_path) as config_json:
		config = json.load(config_json)


	templateData = {
	'config': config
	}
	return render_template('main.html', async_mode=socketio.async_mode, **templateData)

@app.route("/howto")
def howto():
	templateData = {
	'config': config
	}
	return render_template('howto.html', async_mode=socketio.async_mode, **templateData)	

@app.route("/config")
def config():
	with open(config_path) as config_json:
		config = json.load(config_json)

	# max_bot = max(bot['bot_id'] for bot in config['bots'])
	# next_bot = int(max_bot) + 1
	next_bot = helper.get_next_bot(config)

	templateData = {
		'config': config,
		'next_botid': next_bot
	}
	return render_template('config.html', async_mode=socketio.async_mode, **templateData)	
# test websocket actions
# @socketio.on('load_bot', namespace='/test')
# def ping_pong(json):
#     emit('pong', json)

@socketio.on('save_bot',namespace ='/test')
def savebot(input_json):
	#load bot_config.json to object
	with open(config_path) as config_json:
		config = json.load(config_json)
	
	#load json from web socket to object
	new_bot = json.loads(input_json)

	#add and save new object to bot_config.json
	if not any(bot['bot_id'] == new_bot['bot_id'] for bot in config['bots']):
		with open(config_path,'w') as outfile:
			json.dump(helper.add_bot(config,new_bot), outfile)
	else:
		with open(config_path,'w') as outfile:
			json.dump(helper.update_bot(config,new_bot), outfile)

	socketio.emit('confirm_save',new_bot, namespace='/test')

@socketio.on('delete_bot',namespace ='/test')
def deletebot(bot_id):
	with open(config_path) as config_json:
		config = json.load(config_json)

	if any(bot['bot_id'] == bot_id for bot in config['bots']):
		with open(config_path,'w') as outfile:
			json.dump(helper.delete_bot(config,bot_id), outfile)

@socketio.on('load_bot', namespace='/test')
def loadbot(msg):
	with open(config_path) as config_json:
		config = json.load(config_json)

	socketio.emit('bot_info', helper.load_bot(config,msg['bot_id']), namespace='/test')


@socketio.on('startme', namespace='/test')
def startme(msg):
	thread_stop_event.set()
	emit('my_response', {'data': 'Connected', 'count': 0})
	socketio.emit('process_started', {'data': 'Starting Robot', 'time':log_time()}, namespace='/test')
	thread = socketio.start_background_task(runPyBot, msg)
    


@socketio.on('stopme', namespace='/test')
def stopme():
    #socketio.emit('process_stopped',{'data': 'stop that thread'},namespace='/test')
    thread_stop_event.set()

def runPyBot(msg):
	socketio.emit('pong',helper.test_load(msg['bot_id']), namespace='/test')
	bot = helper.test_load(msg['bot_id'])
	for step in bot[0]['step']:
		socketio.emit('pong',step, namespace='/test')
		if step['action'] == 'Typing Bot':
			botgui.typewrite(step['input'][0])

		if step['action'] == 'Keypress Bot':
			botgui.press(step['input'][0])

		if step['action'] == 'Hotkey Bot':
			if len(step['input'][0].split(',')) == 2:
				botgui.hotkey(step['input'][0].split(',')[0],step['input'][0].split(',')[1])
			elif len(step['input'][0].split(',')) == 3:
				botgui.hotkey(step['input'][0].split(',')[0],step['input'][0].split(',')[1],step['input'][0].split(',')[2]) 

		if step['action'] == 'Run Powershell':
			#always reset bot error status to false 
			bot_error = False

			#Run Powershell to aggrigate data - throw error if powershell fails	
			try:
				socketio.emit('pong',step['input'][0], namespace='/test')
				subprocess.check_call([r'C:\\WINDOWS\\system32\\WindowsPowerShell\\v1.0\\powershell.exe',
									'-ExecutionPolicy',
		                            'Unrestricted',
		                            step['input'][0]])
			except subprocess.CalledProcessError:
				bot_error = True

			if bot_error == False:
				socketio.emit('ps_error',{'data': 'Powershell in progress'}, namespace='/test')
			else: #when bot_error == True
				socketio.emit('ps_error',{'data': 'Powershell Error'}, namespace='/test')

		if step['action'] == 'Run Robot Script':
			#build data array
			data_wb = load_workbook(step['input'][1])
			data_ws = data_wb['Sheet1']
			data_array = []
			for cell in data_ws['A']:
				if cell.value != None:
					data_array.append(cell.value)

			socketio.emit('pong',data_array[2], namespace='/test')
			#build action array
			action_wb = load_workbook(step['input'][0])
			action_ws = action_wb['Sheet1']
			action_array = []
			for cell in action_ws['C']:
				if cell.value != None:
					action_array.append(cell.value)

			socketio.emit('pong',action_array[2], namespace='/test')

			dataVar = 0 #setting variable so the script stats at the begining
			wait = .05 #set delay on keyboard input
			num_steps = len(action_array) #count number of steps for use in progress bar
			thread_stop_event.clear() #clear the "stop robot" event
			#completion_time = datetime.datetime.now().strftime("%m-%d-%y %H:%M") #get current time for logging

			#loop through each step of the bot and take action
			for i in range(len(action_array)):
				if not thread_stop_event.isSet():
					if action_array[i] == '$data':
						botgui.typewrite(data_array[dataVar],wait)
						dataVar = dataVar + 1
						socketio.sleep(wait)
					elif action_array[i] == '$save':
						botgui.hotkey('ctrl','s')
						socketio.sleep(wait)
					else:
						botgui.press(action_array[i])
						socketio.sleep(wait)

					#update web page
					number = math.floor((i*100)/num_steps) #calculate pct complete for progress bar
					#send msg to log
					socketio.emit('my_response',{'data': number, 'count': 0}, namespace='/test')
					#send msg to progress bar
					socketio.emit('process_status', {'step': i + 1, 'total_steps': num_steps}, namespace='/test')

			#send process completed or process aborted
			if not thread_stop_event.isSet():
				socketio.emit('my_response',{'data': 100, 'count': 0}, namespace='/test')
				socketio.emit('bot_complete',{'data': 'Process Completed', 'time':log_time()}, namespace='/test')
			else:
				socketio.emit('bot_aborted',{'data': 'Process Aborted', 'time':log_time()}, namespace='/test')
	# else: #when bot_error == True
	# 	socketio.emit('ps_error',{'data': 'Powershell Error'}, namespace='/test')

def log_time():
	return datetime.datetime.now().strftime("%m-%d-%y %H:%M") 

if __name__ == "__main__":
	socketio.run(app,host='0.0.0.0', port=80, debug=True)
